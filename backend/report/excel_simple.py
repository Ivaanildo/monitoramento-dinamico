"""Gerador de relatório Excel e CSV para o Monitoramento Dinâmico.

Aba Visão Geral  — dados de todas as rotas corporativas.
Aba Incidentes   — lista detalhada de incidentes HERE por rota.
Aba Resumo       — resumo executivo com contagens por status/ocorrência/confiança.
"""
import csv
import io
import unicodedata
from datetime import datetime, timedelta, timezone

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

_BRT = timezone(timedelta(hours=-3))

# ===== Paleta de cores =====

CORES = {
    "header_bg":           "0F4C81",
    "header_font":         "FFFFFF",
    "title_bg":            "EAF2FA",
    "subtitle_font":       "4F5D6B",
    "normal_bg":           "D5F5E3",
    "normal_font":         "1E8449",
    "moderado_bg":         "FEF9E7",
    "moderado_font":       "9A7D0A",
    "intenso_bg":          "FDEDEC",
    "intenso_font":        "B03A2E",
    "parado_bg":           "E5E8E8",
    "parado_font":         "1B2631",
    "interdicao_bg":       "F4ECF7",
    "interdicao_font":     "6C3483",
    "bloqueio_parcial_bg": "FDEBD0",
    "bloqueio_parcial_font":"784212",
    "acidente_bg":         "FADBD8",
    "acidente_font":       "922B21",
    "erro_bg":             "E5E7E9",
    "erro_font":           "5D6D7E",
    "zebra_bg":            "F8FAFC",
    "link_font":           "1F5A99",
    "alta_bg":             "D4EFDF",
    "alta_font":           "145A32",
    "media_bg":            "FCF3CF",
    "media_font":          "7D6608",
    "baixa_bg":            "FADBD8",
    "baixa_font":          "922B21",
    "clima_bg":            "DBEAFE",
    "clima_font":          "1E40AF",
    "sem_ocorrencia_bg":   "F3F4F6",
    "sem_ocorrencia_font": "6B7280",
}

HEADER_FILL  = PatternFill("solid", fgColor=CORES["header_bg"])
HEADER_FONT  = Font(name="Calibri", size=11, bold=True, color=CORES["header_font"])
TITLE_FONT   = Font(name="Calibri", size=14, bold=True, color=CORES["header_bg"])
SUBTITLE_FONT= Font(name="Calibri", size=10, italic=True, color=CORES["subtitle_font"])
TITLE_FILL   = PatternFill("solid", fgColor=CORES["title_bg"])
DEFAULT_FONT = Font(name="Calibri", size=10)
LINK_FONT    = Font(name="Calibri", size=10, color=CORES["link_font"], underline="single")
THIN_BORDER  = Border(
    left=Side(style="thin",   color="D5D8DC"),
    right=Side(style="thin",  color="D5D8DC"),
    top=Side(style="thin",    color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ZEBRA_FILL= PatternFill("solid", fgColor=CORES["zebra_bg"])

# ===== Style map =====

_STYLE_MAP = {
    "status": {
        "normal":   ("normal_bg",           "normal_font"),
        "moderado": ("moderado_bg",         "moderado_font"),
        "intenso":  ("intenso_bg",          "intenso_font"),
        "parado":   ("parado_bg",           "parado_font"),
    },
    "ocorrencia": {
        "colisao":           ("acidente_bg",        "acidente_font"),
        "acidente":          ("acidente_bg",        "acidente_font"),
        "bloqueio parcial":  ("bloqueio_parcial_bg","bloqueio_parcial_font"),
        "interdicao":        ("interdicao_bg",      "interdicao_font"),
        "obras na pista":    ("moderado_bg",        "moderado_font"),
        "engarrafamento":    ("intenso_bg",         "intenso_font"),
        "condicao climatica":("clima_bg",           "clima_font"),
        "sem ocorrencia":    ("sem_ocorrencia_bg",  "sem_ocorrencia_font"),
    },
    "confianca": {
        "alta":  ("alta_bg",  "alta_font"),
        "media": ("media_bg", "media_font"),
        "média": ("media_bg", "media_font"),
        "baixa": ("baixa_bg", "baixa_font"),
    },
}


def _norm(txt: str) -> str:
    base = unicodedata.normalize("NFKD", str(txt or ""))
    return "".join(ch for ch in base if not unicodedata.combining(ch)).strip().lower()


def _get_style(categoria: str, valor: str):
    """Retorna (fill, font) para categoria+valor. Fallback por categoria."""
    key = _norm(valor)
    colors = _STYLE_MAP.get(categoria, {}).get(key)
    if colors:
        bg_key, fg_key = colors
        return (
            PatternFill("solid", fgColor=CORES[bg_key]),
            Font(name="Calibri", size=10, bold=True, color=CORES[fg_key]),
        )
    if categoria == "status":
        return (
            PatternFill("solid", fgColor=CORES["erro_bg"]),
            Font(name="Calibri", size=10, bold=True, color=CORES["erro_font"]),
        )
    return None, None


def _texto_curto(texto, limite=320) -> str:
    valor = " ".join(str(texto or "").split())
    if len(valor) <= limite:
        return valor
    return valor[: limite - 3].rstrip() + "..."


def _aplicar_header(ws, headers: list, row: int = 1) -> None:
    for col, (header, width) in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=header)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width


def _aplicar_linha_base(ws, row: int, total_cols: int, zebra: bool = False) -> None:
    for col in range(1, total_cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = DEFAULT_FONT
        c.border = THIN_BORDER
        c.alignment = CENTER
        if zebra:
            c.fill = ZEBRA_FILL


# ===== Excel — consulta individual =====

_HEADERS_DETALHADO = [
    ("Rota",                   40),
    ("Status",                 12),
    ("Atraso (min)",           13),
    ("Confiança",              12),
    ("Incidente Principal",    22),
    ("Vel. Atual (km/h)",      16),
    ("Jam Factor",             12),
    ("Fontes",                 28),
    ("Atualizado em",          20),
    ("Link Waze",              14),
    ("Link Google Maps",       16),
]
TOTAL_COLS_DETALHADO = len(_HEADERS_DETALHADO)


def gerar_excel(resultado: dict) -> bytes:
    """Gera relatório Excel para um ResultadoRota individual. Retorna bytes do .xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Consulta"

    last_col_letter = get_column_letter(TOTAL_COLS_DETALHADO)
    ws.merge_cells(f"A1:{last_col_letter}1")
    t = ws["A1"]
    t.value = f"Monitoramento Dinâmico — Consulta de Rota — {datetime.now(_BRT).strftime('%d/%m/%Y %H:%M')}"
    t.font = TITLE_FONT
    t.fill = TITLE_FILL
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26

    _aplicar_header(ws, headers=_HEADERS_DETALHADO, row=2)

    rota = f"{resultado.get('hub_origem', resultado.get('origem', ''))} → {resultado.get('hub_destino', resultado.get('destino', ''))}"
    status_val = resultado.get("status", "Sem dados")
    atraso     = resultado.get("atraso_min", 0)
    confianca  = resultado.get("confianca", "")
    inc_prin   = resultado.get("incidente_principal") or {}
    inc_label  = inc_prin.get("categoria", "") if isinstance(inc_prin, dict) else str(inc_prin)
    vel_atual  = resultado.get("velocidade_atual_kmh", 0.0)
    jam        = resultado.get("jam_factor_avg", 0.0)
    fontes     = ", ".join(resultado.get("fontes", []))
    consultado = resultado.get("consultado_em", "")
    link_waze  = resultado.get("link_waze", "")
    link_gmaps = resultado.get("link_gmaps", "")

    row = 3
    _aplicar_linha_base(ws, row, total_cols=TOTAL_COLS_DETALHADO)

    valores = [rota, status_val, atraso, confianca, inc_label,
               vel_atual, jam, fontes, consultado, link_waze, link_gmaps]

    for col, val in enumerate(valores, 1):
        c = ws.cell(row=row, column=col, value=val)
        if col == 1:
            c.alignment = LEFT
        if col == 10 and link_waze:
            c.value = "Abrir Waze"
            c.hyperlink = link_waze
            c.font = LINK_FONT
        if col == 11 and link_gmaps:
            c.value = "Abrir Maps"
            c.hyperlink = link_gmaps
            c.font = LINK_FONT

    sf, sft = _get_style("status", status_val)
    ws.cell(row=row, column=2).fill = sf
    ws.cell(row=row, column=2).font = sft

    cf, cft = _get_style("confianca", confianca)
    if cf:
        ws.cell(row=row, column=4).fill = cf
        ws.cell(row=row, column=4).font = cft

    of_, oft = _get_style("ocorrencia", inc_label)
    if of_:
        ws.cell(row=row, column=5).fill = of_
        ws.cell(row=row, column=5).font = oft

    ws.row_dimensions[row].height = 22

    if resultado.get("incidentes"):
        _gerar_aba_incidentes_simples(wb, resultado["incidentes"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _gerar_aba_incidentes_simples(wb: Workbook, incidentes: list) -> None:
    ws = wb.create_sheet("Incidentes HERE")
    inc_headers = [
        ("#", 5), ("Categoria", 18), ("Severidade", 12), ("Rodovia", 16),
        ("Road Closed", 12), ("Descrição", 56), ("Início", 20), ("Fim", 20),
    ]
    for col, (h, w) in enumerate(inc_headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = w

    for idx, inc in enumerate(incidentes, 1):
        row = idx + 1
        for col in range(1, len(inc_headers) + 1):
            c = ws.cell(row=row, column=col)
            c.font = DEFAULT_FONT
            c.border = THIN_BORDER
            c.alignment = CENTER
            if idx % 2 == 0:
                c.fill = ZEBRA_FILL
        vals = [
            idx,
            inc.get("categoria", ""),
            inc.get("severidade", ""),
            inc.get("rodovia_afetada", ""),
            "Sim" if inc.get("road_closed") else "Não",
            inc.get("descricao", ""),
            inc.get("inicio", ""),
            inc.get("fim", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            if col == 6:
                c.alignment = LEFT

    ws.freeze_panes = "A2"


# ===== Excel — Visão Geral (múltiplas rotas) =====

_HEADERS_VISAO_GERAL = [
    ("#",               5),
    ("Sigla",          20),
    ("Nome",           36),
    ("Trecho",         30),
    ("Status",         12),
    ("Ocorrência",     20),
    ("Relato",         50),
    ("Waze",           10),
    ("Google Maps",    14),
    ("Confiança (%)",  14),
    ("Atraso (min)",   13),
    ("Distância (km)", 14),
    ("Atualizado em",  20),
]
TOTAL_COLS_VISAO_GERAL = len(_HEADERS_VISAO_GERAL)


def gerar_excel_visao_geral(resultados: list) -> bytes:
    """Gera relatório Excel para a Visão Geral (múltiplas rotas). Retorna bytes do .xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Visão Geral"

    total     = len(resultados)
    sem_dados = sum(1 for r in resultados if r.get("status") in ("Erro", "N/A", "Sem dados"))
    last_col  = get_column_letter(TOTAL_COLS_VISAO_GERAL)

    # Título
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value = f"Monitoramento Dinâmico — Visão Geral Analítica — {datetime.now(_BRT).strftime('%d/%m/%Y %H:%M')}"
    t.font = TITLE_FONT
    t.fill = TITLE_FILL
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26

    # Subtítulo
    ws.merge_cells(f"A2:{last_col}2")
    s = ws["A2"]
    s.value = "Dados consolidados das rotas corporativas monitoradas"
    s.font = SUBTITLE_FONT
    s.alignment = CENTER

    # Aviso de cobertura baixa (>50% sem dados)
    header_row = 4
    if total > 0 and sem_dados / total > 0.5:
        ws.merge_cells(f"A3:{last_col}3")
        aviso = ws["A3"]
        aviso.value = f"ATENÇÃO: {sem_dados}/{total} rotas sem dados. Verifique as API keys."
        aviso.font = Font(name="Calibri", size=10, bold=True, color=CORES["intenso_font"])
        aviso.fill = PatternFill("solid", fgColor=CORES["intenso_bg"])
        aviso.alignment = CENTER
        header_row = 5

    _aplicar_header(ws, headers=_HEADERS_VISAO_GERAL, row=header_row)

    for idx, r in enumerate(resultados, 1):
        row = idx + header_row
        _aplicar_linha_base(ws, row, TOTAL_COLS_VISAO_GERAL, zebra=(idx % 2 == 0))

        status_val   = r.get("status", "Sem dados")
        ocorrencia   = r.get("ocorrencia", "")
        relato       = r.get("relato", "")
        confianca_pct= r.get("confianca_pct", 0)
        atraso       = r.get("atraso_min", 0)
        distancia    = r.get("distancia_km", 0.0)
        link_waze    = r.get("link_waze", "")
        link_gmaps   = r.get("link_gmaps", "")
        atualizado   = r.get("hora_atualizacao", "")

        valores = [
            idx,
            r.get("sigla", ""),
            r.get("nome", ""),
            r.get("trecho", ""),
            status_val,
            ocorrencia,
            _texto_curto(relato),
            link_waze,
            link_gmaps,
            confianca_pct,
            atraso,
            distancia,
            atualizado,
        ]

        for col, val in enumerate(valores, 1):
            c = ws.cell(row=row, column=col, value=val)
            if col in (3, 4, 7):
                c.alignment = LEFT
            if col == 8 and link_waze:
                c.value = "Abrir Waze"
                c.hyperlink = link_waze
                c.font = LINK_FONT
            if col == 9 and link_gmaps:
                c.value = "Abrir Maps"
                c.hyperlink = link_gmaps
                c.font = LINK_FONT

        # Estilos por valor
        sf, sft = _get_style("status", status_val)
        ws.cell(row=row, column=5).fill = sf
        ws.cell(row=row, column=5).font = sft

        of_, oft = _get_style("ocorrencia", ocorrencia)
        if of_:
            ws.cell(row=row, column=6).fill = of_
            ws.cell(row=row, column=6).font = oft

        if confianca_pct >= 90:
            str_conf = "alta"
        elif confianca_pct >= 50:
            str_conf = "media"
        else:
            str_conf = "baixa"
        cf, cft = _get_style("confianca", str_conf)
        if cf:
            ws.cell(row=row, column=10).fill = cf
            ws.cell(row=row, column=10).font = cft

        # Altura dinâmica baseada no conteúdo
        max_len = max(len(str(relato or "")), len(str(ocorrencia or "")))
        ws.row_dimensions[row].height = min(90, 22 + (max_len // 60) * 8)

    # Legenda
    legenda_start = header_row + total + 3
    ws.cell(row=legenda_start, column=1, value="Legenda").font = Font(name="Calibri", size=10, bold=True)
    legendas = [
        ("Status Normal",               CORES["normal_bg"],           CORES["normal_font"]),
        ("Status Moderado",             CORES["moderado_bg"],         CORES["moderado_font"]),
        ("Status Intenso",              CORES["intenso_bg"],          CORES["intenso_font"]),
        ("Status Parado",               CORES["parado_bg"],           CORES["parado_font"]),
        ("Ocorrência Colisão",          CORES["acidente_bg"],         CORES["acidente_font"]),
        ("Ocorrência Interdição",       CORES["interdicao_bg"],       CORES["interdicao_font"]),
        ("Ocorrência Bloqueio Parcial", CORES["bloqueio_parcial_bg"], CORES["bloqueio_parcial_font"]),
    ]
    for offset, (label, bg, fg) in enumerate(legendas, 1):
        c = ws.cell(row=legenda_start + offset, column=1, value=label)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Calibri", size=9, bold=True, color=fg)
        c.border = THIN_BORDER

    # Auto-filter e freeze
    max_data_row = header_row + total
    ws.auto_filter.ref = f"A{header_row}:{last_col}{max_data_row}"
    ws.freeze_panes = f"A{header_row + 1}"

    # Formatação condicional — Status (col E = 5)
    col_status = get_column_letter(5)
    range_status = f"{col_status}{header_row + 1}:{col_status}{max_data_row}"
    dv_status = DataValidation(
        type="list",
        formula1='"Normal,Moderado,Intenso,Parado"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_status)
    dv_status.add(range_status)
    for texto, (bg_key, font_key) in [
        ("Normal",   ("normal_bg",   "normal_font")),
        ("Moderado", ("moderado_bg", "moderado_font")),
        ("Intenso",  ("intenso_bg",  "intenso_font")),
        ("Parado",   ("parado_bg",   "parado_font")),
    ]:
        ws.conditional_formatting.add(
            range_status,
            CellIsRule(
                operator="equal",
                formula=[f'"{texto}"'],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor=CORES[bg_key]),
                font=Font(name="Calibri", size=10, bold=True, color=CORES[font_key]),
            ),
        )

    # Abas adicionais
    _gerar_aba_incidentes_visao_geral(wb, resultados)
    _gerar_aba_resumo(wb, resultados)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _gerar_aba_incidentes_visao_geral(wb: Workbook, resultados: list) -> None:
    ws = wb.create_sheet("Incidentes")
    headers = [
        ("#",           5),
        ("Rota",       28),
        ("Categoria",  18),
        ("Severidade", 12),
        ("Rodovia",    16),
        ("Road Closed",12),
        ("Descrição",  56),
        ("Início",     20),
        ("Fim",        20),
    ]
    _aplicar_header(ws, headers, row=1)

    linhas = []
    for r in resultados:
        rota_nome = r.get("nome", r.get("sigla", ""))
        for inc in r.get("incidentes", []):
            linhas.append([
                rota_nome,
                inc.get("categoria", ""),
                inc.get("severidade", ""),
                inc.get("rodovia_afetada", ""),
                "Sim" if inc.get("road_closed") else "Não",
                inc.get("descricao", ""),
                inc.get("inicio", ""),
                inc.get("fim", ""),
            ])

    if not linhas:
        ws.cell(row=2, column=1, value="Nenhum incidente registrado neste ciclo.").font = Font(
            name="Calibri", size=10, italic=True
        )
    else:
        for idx, linha in enumerate(linhas, 1):
            row = idx + 1
            _aplicar_linha_base(ws, row, len(headers), zebra=(idx % 2 == 0))
            ws.cell(row=row, column=1, value=idx)
            for col, val in enumerate(linha, 2):
                c = ws.cell(row=row, column=col, value=val)
                c.alignment = LEFT if col in (2, 7) else CENTER

    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{max(1, len(linhas) + 1)}"
    ws.freeze_panes = "A2"


def _gerar_aba_resumo(wb: Workbook, resultados: list) -> None:
    ws = wb.create_sheet("Resumo")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14

    ws["A1"] = "Resumo Executivo"
    ws["A1"].font = TITLE_FONT

    ws["A3"] = "Data/Hora"
    ws["A3"].font = HEADER_FONT
    ws["A3"].fill = HEADER_FILL
    ws["B3"] = datetime.now(_BRT).strftime("%d/%m/%Y %H:%M:%S")

    ws["A4"] = "Rotas monitoradas"
    ws["A4"].font = HEADER_FONT
    ws["A4"].fill = HEADER_FILL
    ws["B4"] = len(resultados)

    status_count: dict    = {}
    ocorrencia_count: dict = {}
    confianca_count: dict  = {}

    for r in resultados:
        status    = r.get("status", "Sem dados")
        ocorrencia= (r.get("ocorrencia", "") or "").strip()
        pct       = r.get("confianca_pct", 0) or 0

        status_count[status] = status_count.get(status, 0) + 1

        chave_occ = ocorrencia if ocorrencia else "Sem ocorrência"
        ocorrencia_count[chave_occ] = ocorrencia_count.get(chave_occ, 0) + 1

        nivel = "Alta" if pct >= 90 else ("Média" if pct >= 50 else "Baixa")
        confianca_count[nivel] = confianca_count.get(nivel, 0) + 1

    row = 6
    ws.cell(row=row, column=1, value="Status").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    row += 1
    for st, qty in sorted(status_count.items(), key=lambda x: (-x[1], x[0])):
        ws.cell(row=row, column=1, value=st).border  = THIN_BORDER
        ws.cell(row=row, column=2, value=qty).border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Ocorrências").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    row += 1
    for occ, qty in sorted(ocorrencia_count.items(), key=lambda x: (-x[1], x[0])):
        ws.cell(row=row, column=1, value=occ).border  = THIN_BORDER
        ws.cell(row=row, column=2, value=qty).border  = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Confiança").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    row += 1
    for nivel, qty in sorted(confianca_count.items(), key=lambda x: (-x[1], x[0])):
        ws.cell(row=row, column=1, value=nivel).border = THIN_BORDER
        ws.cell(row=row, column=2, value=qty).border   = THIN_BORDER
        row += 1


# ===== CSV =====

def gerar_csv(resultado: dict) -> str:
    """Gera relatório CSV simples para um ResultadoRota. Retorna string."""
    buf = io.StringIO()
    writer = csv.writer(buf)

    writer.writerow([
        "Rota", "Status", "Atraso (min)", "Confiança", "Incidente Principal",
        "Vel. Atual (km/h)", "Jam Factor Avg", "Jam Factor Max",
        "Distância (km)", "Duração Normal (min)", "Duração Tráfego (min)",
        "Razão", "Fontes", "Atualizado em", "Link Waze", "Link Google Maps",
    ])

    rota = f"{resultado.get('hub_origem', resultado.get('origem', ''))} → {resultado.get('hub_destino', resultado.get('destino', ''))}"
    inc_prin = resultado.get("incidente_principal") or {}
    inc_label = inc_prin.get("categoria", "") if isinstance(inc_prin, dict) else str(inc_prin)

    writer.writerow([
        rota,
        resultado.get("status", ""),
        resultado.get("atraso_min", 0),
        resultado.get("confianca", ""),
        inc_label,
        resultado.get("velocidade_atual_kmh", ""),
        resultado.get("jam_factor_avg", ""),
        resultado.get("jam_factor_max", ""),
        resultado.get("distancia_km", ""),
        resultado.get("duracao_normal_min", ""),
        resultado.get("duracao_transito_min", ""),
        resultado.get("razao_transito", ""),
        ", ".join(resultado.get("fontes", [])),
        resultado.get("consultado_em", ""),
        resultado.get("link_waze", ""),
        resultado.get("link_gmaps", ""),
    ])

    if resultado.get("incidentes"):
        writer.writerow([])
        writer.writerow(["=== Incidentes HERE ==="])
        writer.writerow(["Categoria", "Severidade", "Rodovia", "Road Closed", "Descrição", "Início", "Fim"])
        for inc in resultado["incidentes"]:
            writer.writerow([
                inc.get("categoria", ""),
                inc.get("severidade", ""),
                inc.get("rodovia_afetada", ""),
                "Sim" if inc.get("road_closed") else "Não",
                inc.get("descricao", ""),
                inc.get("inicio", ""),
                inc.get("fim", ""),
            ])

    return buf.getvalue()


def gerar_csv_visao_geral(resultados: list) -> str:
    """Gera relatório CSV para a Visão Geral (múltiplas rotas). Retorna string."""
    buf = io.StringIO()
    writer = csv.writer(buf)

    writer.writerow([
        "ID", "Sigla", "Nome", "Trecho", "Status",
        "Ocorrência", "Relato", "Atualização", "Confiança (%)",
        "Atraso (min)", "Distância (km)", "Link Waze", "Link Google Maps",
    ])

    for r in resultados:
        writer.writerow([
            r.get("rota_id", ""),
            r.get("sigla", ""),
            r.get("nome", ""),
            r.get("trecho", ""),
            r.get("status", ""),
            r.get("ocorrencia", ""),
            r.get("relato", ""),
            r.get("hora_atualizacao", ""),
            r.get("confianca_pct", ""),
            r.get("atraso_min", ""),
            r.get("distancia_km", ""),
            r.get("link_waze", ""),
            r.get("link_gmaps", ""),
        ])

    return buf.getvalue()
